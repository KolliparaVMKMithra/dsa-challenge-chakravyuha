import datetime
import os
import io
import logging
from fastapi import APIRouter, Depends, HTTPException, status, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from backend.database import get_db
from backend.models import Student, Problem, Submission, Attendance, CodeChefContest, CodeChefParticipation, Feedback, Event, EventRegistration, SIHTeam, SIHTeamMember, SIHProblemStatement, SIHPSSelection, SIHJudgingScore
from backend.schemas import ProblemCreate, ProblemResponse, CodeChefContestCreate, CodeChefContestResponse, ScanAdminCreate, ScanAdminResponse, EventCreate, EventResponse, SIHTeamRegistration, AdminPSOverrideRequest
from backend.auth import get_current_attendance_admin, get_current_super_admin, get_password_hash

router = APIRouter(prefix="/api/admin", tags=["admin"])
logger = logging.getLogger(__name__)

DEBUG_EMAILS_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "debug_emails.log")

@router.post("/scan")
def scan_qr(payload: Dict[str, Any], current_admin: Student = Depends(get_current_attendance_admin), db: Session = Depends(get_db)):
    """Scans a student's QR key and marks them present for the day or for a specific event."""
    qr_key = payload.get("qr_key")
    session = payload.get("session", "forenoon")
    event_id = payload.get("event_id")
    
    if not qr_key:
        raise HTTPException(status_code=400, detail="qr_key is required")
        
    student = db.query(Student).filter(Student.qr_key == qr_key).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found with this QR code")
        
    # Detect if session starts with "event_"
    if not event_id and isinstance(session, str) and session.startswith("event_"):
        try:
            event_id = int(session.split("_")[1])
        except ValueError:
            pass

    if event_id is not None:
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found.")
            
        reg = db.query(EventRegistration).filter(
            EventRegistration.student_id == student.id,
            EventRegistration.event_id == event_id
        ).first()
        
        if not reg:
            raise HTTPException(
                status_code=400,
                detail=f"{student.full_name} ({student.roll_number}) is not registered for event {event.name}."
            )
            
        if reg.attended:
            raise HTTPException(
                status_code=400,
                detail=f"Attendance already marked for {student.full_name} ({student.roll_number}) for event {event.name}."
            )
            
        reg.attended = True
        reg.attended_at = datetime.datetime.utcnow()
        reg.attendance_marked_by = current_admin.full_name
        db.commit()
        db.refresh(reg)
        
        return {
            "success": True,
            "student_name": student.full_name,
            "name": student.full_name,
            "roll_number": student.roll_number,
            "branch": student.branch,
            "year": student.year,
            "session": f"Event: {event.name}",
            "time": reg.attended_at.strftime("%d-%m-%Y %I:%M %p")
        }
        
    today = datetime.date.today()
    
    # Check duplicate daily attendance
    existing = db.query(Attendance).filter(
        Attendance.student_id == student.id,
        Attendance.date == today,
        Attendance.session == session
    ).first()
    
    if existing:
        raise HTTPException(
            status_code=400, 
            detail=f"Attendance already marked for {str(session).capitalize()} today for {student.full_name} ({student.roll_number})"
        )
        
    # Mark present
    attendance = Attendance(
        student_id=student.id,
        date=today,
        session=session,
        marked_by=current_admin.full_name
    )
    db.add(attendance)
    db.commit()
    db.refresh(attendance)
    
    return {
        "success": True,
        "student_name": student.full_name,
        "name": student.full_name,
        "roll_number": student.roll_number,
        "branch": student.branch,
        "year": student.year,
        "session": str(session).capitalize(),
        "time": attendance.timestamp.strftime("%d-%m-%Y %I:%M %p")
    }

@router.get("/attendance/today")
def get_today_attendance(session: str = "forenoon", event_id: Optional[str] = None, current_admin: Student = Depends(get_current_attendance_admin), db: Session = Depends(get_db)):
    """Lists all students marked present today for the specified session or event."""
    parsed_event_id: Optional[int] = None
    
    if event_id is not None:
        val = str(event_id).strip()
        if val and val.lower() not in ("null", "undefined"):
            try:
                parsed_event_id = int(val)
            except ValueError:
                raise HTTPException(status_code=400, detail="event_id must be a valid integer")
                
    if not parsed_event_id and isinstance(session, str) and session.startswith("event_"):
        try:
            parsed_event_id = int(session.split("_")[1])
        except ValueError:
            pass

    if parsed_event_id is not None:
        regs = db.query(EventRegistration).filter(
            EventRegistration.event_id == parsed_event_id,
            EventRegistration.attended == True
        ).all()
        
        result = []
        for r in regs:
            timestamp_val = r.attended_at if r.attended_at else r.registered_at
            result.append({
                "timestamp": timestamp_val,
                "marked_by": r.attendance_marked_by or "System",
                "full_name": r.student.full_name,
                "name": r.student.full_name,
                "roll_number": r.student.roll_number or "N/A",
                "branch": r.student.branch,
                "year": r.student.year,
                "session": "Event Attendance"
            })
        result.sort(key=lambda x: x["timestamp"], reverse=True)
        return result

    today = datetime.date.today()
    records = db.query(
        Attendance.timestamp,
        Attendance.marked_by,
        Student.full_name,
        Student.roll_number,
        Student.branch,
        Student.year,
        Attendance.session
    ).join(Student, Attendance.student_id == Student.id)\
     .filter(Attendance.date == today, Attendance.session == session)\
     .order_by(Attendance.timestamp.desc()).all()
     
    return [
        {
            "timestamp": r[0],
            "marked_by": r[1],
            "full_name": r[2],
            "name": r[2],
            "roll_number": r[3],
            "branch": r[4],
            "year": r[5],
            "session": r[6]
        } for r in records
    ]


@router.get("/attendance/export")
def export_attendance(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    current_admin: Student = Depends(get_current_attendance_admin),
    db: Session = Depends(get_db)
):
    """Exports attendance log as a styled Excel sheet using openpyxl."""
    query = db.query(
        Attendance.date,
        Attendance.timestamp,
        Attendance.marked_by,
        Student.full_name,
        Student.roll_number,
        Student.branch,
        Student.year,
        Attendance.session
    ).join(Student, Attendance.student_id == Student.id)
    
    if start_date:
        try:
            s_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
            query = query.filter(Attendance.date >= s_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid start_date format, use YYYY-MM-DD")
            
    if end_date:
        try:
            e_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
            query = query.filter(Attendance.date <= e_date)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid end_date format, use YYYY-MM-DD")
            
    records = query.order_by(Attendance.date.desc(), Attendance.timestamp.desc()).all()
    
    # Generate Excel using openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Logs"
    
    # Title Banner
    ws.merge_cells("A1:H1")
    ws["A1"] = "Chakravyuha Daily DSA Challenge - Attendance Logs"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="8C7030", end_color="8C7030", fill_type="solid") # Gold/amber
    ws.row_dimensions[1].height = 40
    
    # Headers
    headers = ["Date", "Session", "Time", "Roll Number", "Full Name", "Branch", "Year", "Marked By"]
    ws.append([]) # Blank row 2
    ws.append(headers) # Row 3
    
    # Format Headers (Row 3)
    header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    # Append Data
    for rec in records:
        row_data = [
            rec[0].strftime("%Y-%m-%d"),
            rec[7].capitalize(),
            rec[1].strftime("%I:%M %p"),
            rec[4],
            rec[3],
            rec[5],
            rec[6],
            rec[2]
        ]
        ws.append(row_data)
        
    # Style Data Cells
    for row in range(4, ws.max_row + 1):
        # Alternate row fill
        row_fill = PatternFill(start_color="F9F9F9" if row % 2 == 0 else "FFFFFF", end_color="F9F9F9" if row % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in [1, 2, 3, 4, 7]:
                cell.alignment = Alignment(horizontal="center")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                continue
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"attendance_report_{start_date or 'all'}_to_{end_date or 'all'}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/students")
def get_students_directory(
    search: Optional[str] = Query(None),
    branch: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Fetches a detailed list of students with progress summaries."""
    query = db.query(Student).filter(Student.is_admin == False)
    
    if search:
        search_lower = search.lower()
        query = query.filter(
            (func.lower(Student.full_name).like(f"%{search_lower}%")) |
            (func.lower(Student.roll_number).like(f"%{search_lower}%")) |
            (func.lower(Student.college_email).like(f"%{search_lower}%"))
        )
    if branch:
        query = query.filter(Student.branch == branch)
    if year:
        query = query.filter(Student.year == year)
        
    students = query.order_by(Student.roll_number).all()
    
    result = []
    # Fetch problem count for percentage calculations
    total_problems = db.query(func.count(Problem.id)).filter(Problem.is_active == True).scalar() or 0
    
    for s in students:
        solved_count = db.query(func.count(Submission.id)).filter(
            Submission.student_id == s.id,
            Submission.solved == True
        ).scalar() or 0
        
        attendance_count = db.query(func.count(Attendance.id)).filter(Attendance.student_id == s.id).scalar() or 0
        
        result.append({
            "id": s.id,
            "name": s.full_name,
            "roll_number": s.roll_number,
            "email": s.college_email,
            "phone": s.phone_number,
            "branch": s.branch,
            "year": s.year,
            "streak": s.streak_count,
            "solved": solved_count,
            "total_problems": total_problems,
            "percentage": round(solved_count / total_problems * 100, 1) if total_problems > 0 else 0,
            "attendance_count": attendance_count
        })
        
    return result

@router.get("/students/export")
def export_students_directory(
    search: Optional[str] = Query(None),
    branch: Optional[str] = Query(None),
    year: Optional[int] = Query(None),
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Exports the student directory progress report as a styled Excel sheet using openpyxl."""
    query = db.query(Student).filter(Student.is_admin == False)
    if search:
        search_lower = search.lower()
        query = query.filter(
            (func.lower(Student.full_name).like(f"%{search_lower}%")) |
            (func.lower(Student.roll_number).like(f"%{search_lower}%")) |
            (func.lower(Student.college_email).like(f"%{search_lower}%"))
        )
    if branch:
        query = query.filter(Student.branch == branch)
    if year:
        query = query.filter(Student.year == year)
        
    students = query.order_by(Student.roll_number).all()
    total_problems = db.query(func.count(Problem.id)).filter(Problem.is_active == True).scalar() or 0
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Progress"
    
    # Title Banner
    ws.merge_cells("A1:J1")
    ws["A1"] = "Chakravyuha Daily DSA Challenge - Student Progress Report"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="8C7030", end_color="8C7030", fill_type="solid")
    ws.row_dimensions[1].height = 40
    
    headers = ["Roll Number", "Full Name", "Email", "Phone", "Branch", "Year", "Streak", "Forenoon Attendance", "Afternoon Attendance", "Solved", "Total", "Solve %"]
    ws.append([]) # Blank row 2
    ws.append(headers) # Row 3
    
    header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for s in students:
        solved_count = db.query(func.count(Submission.id)).filter(
            Submission.student_id == s.id,
            Submission.solved == True
        ).scalar() or 0
        
        forenoon_count = db.query(func.count(Attendance.id)).filter(
            Attendance.student_id == s.id,
            Attendance.session == "forenoon"
        ).scalar() or 0
        
        afternoon_count = db.query(func.count(Attendance.id)).filter(
            Attendance.student_id == s.id,
            Attendance.session == "afternoon"
        ).scalar() or 0
        
        pct = round(solved_count / total_problems * 100, 1) if total_problems > 0 else 0
        
        row_data = [
            s.roll_number,
            s.full_name,
            s.college_email,
            s.phone_number,
            s.branch,
            s.year,
            s.streak_count,
            forenoon_count,
            afternoon_count,
            solved_count,
            total_problems,
            f"{pct}%"
        ]
        ws.append(row_data)
        
    for row in range(4, ws.max_row + 1):
        row_fill = PatternFill(start_color="F9F9F9" if row % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, 13):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in [1, 5, 6, 7, 8, 9, 10, 11, 12]:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or '')
            if cell.row == 1:
                continue
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_progress_report.xlsx"}
    )

@router.get("/students/{student_id}/detail")
def get_student_detail(student_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Fetches full progress details, submissions, and logs for a specific student."""
    student = db.query(Student).filter(Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
        
    # Submissions
    submissions = db.query(
        Submission.completed_at,
        Submission.submission_link,
        Problem.title,
        Problem.topic,
        Problem.difficulty
    ).join(Problem, Submission.problem_id == Problem.id)\
     .filter(Submission.student_id == student_id).order_by(Submission.completed_at.desc()).all()
     
    # Attendance
    attendance = db.query(Attendance).filter(Attendance.student_id == student_id).order_by(Attendance.date.desc()).all()
    
    # CodeChef
    codechef = db.query(
        CodeChefContest.week_number,
        CodeChefParticipation.status,
        CodeChefParticipation.submission_proof,
        CodeChefParticipation.updated_at
    ).join(CodeChefParticipation, CodeChefContest.id == CodeChefParticipation.contest_id)\
     .filter(CodeChefParticipation.student_id == student_id).order_by(CodeChefContest.week_number.desc()).all()
     
    return {
        "student": {
            "id": student.id,
            "name": student.full_name,
            "roll_number": student.roll_number,
            "email": student.college_email,
            "phone": student.phone_number,
            "branch": student.branch,
            "year": student.year,
            "streak": student.streak_count,
            "qr_key": student.qr_key
        },
        "submissions": [
            {
                "title": s[2],
                "topic": s[3],
                "difficulty": s[4],
                "link": s[1],
                "date": f"{s[0].isoformat()}Z" if (s[0] and s[0] >= datetime.datetime(2026, 7, 10, 15, 0, 0)) else None
            } for s in submissions
        ],
        "attendance": [
            {
                "date": f"{d_str} ({', '.join(['FN' if s == 'forenoon' else 'AN' for s in sorted(list(set(d_data['sessions'])))])})",
                "timestamp": d_data["timestamp"],
                "marked_by": ", ".join(list(set(d_data["marked_by_list"])))
            } for d_str, d_data in (lambda atts: (
                lambda grouped: (
                    [grouped.setdefault(a.date.strftime("%Y-%m-%d"), {"sessions": [], "marked_by_list": [], "timestamp": a.timestamp})["sessions"].append(a.session) or 
                     grouped[a.date.strftime("%Y-%m-%d")]["marked_by_list"].append(a.marked_by)
                     for a in atts],
                    grouped
                )[1]
            )({}))(attendance).items()
        ],
        "codechef": [
            {
                "week": c[0],
                "status": c[1],
                "proof": c[2],
                "date": c[3]
            } for c in codechef
        ]
    }

# ----------------- PROBLEM CRUD -----------------

@router.get("/problems", response_model=List[ProblemResponse])
def list_problems(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    return db.query(Problem).order_by(Problem.topic, Problem.id).all()

@router.post("/problems", response_model=ProblemResponse)
def add_problem(prob: ProblemCreate, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    db_prob = Problem(**prob.model_dump())
    db.add(db_prob)
    db.commit()
    db.refresh(db_prob)
    return db_prob

@router.put("/problems/{problem_id}", response_model=ProblemResponse)
def update_problem(
    problem_id: int, 
    prob_data: ProblemCreate, 
    current_admin: Student = Depends(get_current_super_admin), 
    db: Session = Depends(get_db)
):
    db_prob = db.query(Problem).filter(Problem.id == problem_id).first()
    if not db_prob:
        raise HTTPException(status_code=404, detail="Problem not found")
        
    for k, v in prob_data.model_dump().items():
        setattr(db_prob, k, v)
        
    db.commit()
    db.refresh(db_prob)
    return db_prob

@router.delete("/problems/{problem_id}")
def delete_problem(problem_id: int, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    db_prob = db.query(Problem).filter(Problem.id == problem_id).first()
    if not db_prob:
        raise HTTPException(status_code=404, detail="Problem not found")
    
    # Hard or soft delete? Let's delete
    db.delete(db_prob)
    db.commit()
    return {"detail": "Problem deleted successfully"}

# ----------------- CODECHEF MANAGEMENT -----------------

@router.post("/codechef/contest", response_model=CodeChefContestResponse)
def create_codechef_contest(
    contest_data: CodeChefContestCreate, 
    current_admin: Student = Depends(get_current_super_admin), 
    db: Session = Depends(get_db)
):
    # Check if week already exists
    existing = db.query(CodeChefContest).filter(CodeChefContest.week_number == contest_data.week_number).first()
    if existing:
        # Update it
        existing.contest_link = contest_data.contest_link
        existing.deadline = contest_data.deadline
        db.commit()
        db.refresh(existing)
        return existing
        
    new_contest = CodeChefContest(**contest_data.model_dump())
    db.add(new_contest)
    db.commit()
    db.refresh(new_contest)
    return new_contest

# ----------------- BULK EMAIL BROADCASTER -----------------

@router.post("/bulk-email")
def send_bulk_email(payload: Dict[str, Any], current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Simulates broadcasting emails to filtered or specific students via Office 365 / Power Automate integration."""
    subject = payload.get("subject")
    body = payload.get("body")
    filter_type = payload.get("filter_type", "all") # all, missed_codechef, inactive, or custom
    student_ids = payload.get("student_ids") # list of UUID strings
    
    if not subject or not body:
        raise HTTPException(status_code=400, detail="subject and body are required")
        
    if student_ids is not None:
        recipients = db.query(Student).filter(Student.is_admin == False, Student.id.in_(student_ids)).all()
    else:
        # Get filtered students
        query = db.query(Student).filter(Student.is_admin == False)
        
        if filter_type == "missed_codechef":
            # Find latest CodeChef contest
            latest_contest = db.query(CodeChefContest).order_by(CodeChefContest.week_number.desc()).first()
            if latest_contest:
                # Subquery of students who attended
                attended_sub = db.query(CodeChefParticipation.student_id).filter(
                    CodeChefParticipation.contest_id == latest_contest.id,
                    CodeChefParticipation.status == "attended"
                )
                query = query.filter(~Student.id.in_(attended_sub))
                
        elif filter_type == "inactive":
            # Students with 0 solved problems
            active_sub = db.query(Submission.student_id).filter(Submission.solved == True).distinct()
            query = query.filter(~Student.id.in_(active_sub))
            
        recipients = query.all()
    
    # Format and save debug email log
    log_content = (
        f"========================================\n"
        f"BULK EMAIL BROADCAST: {subject}\n"
        f"Timestamp: {datetime.datetime.utcnow()}\n"
        f"Sender: {current_admin.college_email} (Admin)\n"
        f"Filter Group: {filter_type}\n"
        f"Recipient Count: {len(recipients)}\n"
        f"Recipients: {', '.join([s.college_email for s in recipients])}\n"
        f"----------------------------------------\n"
        f"Body:\n{body}\n"
        f"========================================\n\n"
    )
    
    try:
        with open(DEBUG_EMAILS_PATH, "a", encoding="utf-8") as f:
            f.write(log_content)
    except Exception as e:
        logger.error(f"Failed to log bulk email: {e}")
        
    logger.info(f"Broadcasted bulk email to {len(recipients)} users (filter: {filter_type})")
    
    # Power Automate Webhook Stub
    # If POWER_AUTOMATE_WEBHOOK_URL env is set:
    # requests.post(WEBHOOK_URL, json={"subject": subject, "body": body, "emails": [s.college_email for s in recipients]})
    
    return {
        "success": True,
        "recipient_count": len(recipients),
        "recipients": [s.full_name for s in recipients]
    }

# ----------------- REPORTS & ANALYTICS -----------------

@router.get("/reports/dashboard")
def get_reports_dashboard(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Fetches high-level metrics, leaderboard, solve rates, and CodeChef compliance rates."""
    # 1. High level aggregates
    total_students = db.query(func.count(Student.id)).filter(Student.is_admin == False).scalar() or 0
    total_problems = db.query(func.count(Problem.id)).filter(Problem.is_active == True).scalar() or 0
    
    # 2. Leaderboard (Top 10 solvers)
    leaderboard_query = db.query(
        Student.full_name,
        Student.roll_number,
        Student.branch,
        Student.year,
        Student.streak_count,
        func.count(Submission.id).label("solved_count")
    ).join(Submission, Student.id == Submission.student_id)\
     .filter(Student.is_admin == False, Submission.solved == True)\
     .group_by(Student.id)\
     .order_by(func.count(Submission.id).desc(), Student.streak_count.desc())\
     .limit(10).all()
     
    leaderboard = [
        {
            "name": r[0],
            "roll_number": r[1],
            "branch": r[2],
            "year": r[3],
            "streak": r[4],
            "solved": r[5]
        } for r in leaderboard_query
    ]
    
    # 3. Topic-wise solve rates across the club
    # Number of students solved each topic
    topic_rates = db.query(
        Problem.topic,
        func.count(Submission.id).label("total_solved")
    ).join(Submission, Problem.id == Submission.problem_id)\
     .filter(Submission.solved == True)\
     .group_by(Problem.topic).all()
     
    # Active problems count per topic
    topic_problems_count = db.query(
        Problem.topic,
        func.count(Problem.id)
    ).filter(Problem.is_active == True)\
     .group_by(Problem.topic).all()
     
    prob_counts = {t[0]: t[1] for t in topic_problems_count}
    
    topic_solve_stats = []
    for topic, solved in topic_rates:
        prob_count = prob_counts.get(topic, 0)
        max_possible_solved = total_students * prob_count if total_students > 0 else 0
        solve_rate = (solved / max_possible_solved * 100) if max_possible_solved > 0 else 0
        topic_solve_stats.append({
            "topic": topic,
            "solved_count": solved,
            "total_problems": prob_count,
            "rate": round(solve_rate, 1)
        })
        
    # Sort by rate descending
    topic_solve_stats.sort(key=lambda x: x["rate"], reverse=True)
    
    # 4. CodeChef compliance rates
    # Get latest contest
    latest_contest = db.query(CodeChefContest).order_by(CodeChefContest.week_number.desc()).first()
    compliance = {"week": None, "attended": 0, "missed": 0, "rate": 0}
    
    if latest_contest:
        attended = db.query(func.count(CodeChefParticipation.id)).filter(
            CodeChefParticipation.contest_id == latest_contest.id,
            CodeChefParticipation.status == "attended"
        ).scalar() or 0
        
        missed = total_students - attended
        compliance = {
            "week": latest_contest.week_number,
            "attended": attended,
            "missed": max(0, missed),
            "rate": round(attended / total_students * 100, 1) if total_students > 0 else 0
        }
        
    # 5. Weekly attendance trend (Last 7 days)
    today = datetime.date.today()
    attendance_trend = []
    for i in range(6, -1, -1):
        day = today - datetime.timedelta(days=i)
        att_count = db.query(func.count(Attendance.id)).filter(Attendance.date == day).scalar() or 0
        attendance_trend.append({
            "date": day.strftime("%b %d"),
            "present": att_count,
            "absent": max(0, total_students - att_count)
        })
        
    return {
        "total_students": total_students,
        "total_problems": total_problems,
        "leaderboard": leaderboard,
        "topic_solve_rates": topic_solve_stats,
        "codechef_compliance": compliance,
        "attendance_trend": attendance_trend
    }

# ----------------- SCAN ADMINS MANAGEMENT -----------------

@router.get("/scan-admins", response_model=List[ScanAdminResponse])
def list_scan_admins(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Lists all Scan Admins (attendance role)."""
    return db.query(Student).filter(Student.is_admin == True, Student.admin_role == "attendance").order_by(Student.created_at.desc()).all()

@router.post("/scan-admins", response_model=ScanAdminResponse)
def create_scan_admin(admin_data: ScanAdminCreate, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Creates a new Scan Admin or promotes/updates an existing user."""
    existing = db.query(Student).filter(
        (func.lower(Student.college_email) == func.lower(admin_data.college_email)) |
        (func.lower(Student.roll_number) == func.lower(admin_data.roll_number))
    ).first()

    if existing:
        # Promote or update existing user
        existing.is_admin = True
        existing.admin_role = "attendance"
        existing.full_name = admin_data.full_name
        existing.phone_number = admin_data.phone_number
        if admin_data.password:
            existing.password_hash = get_password_hash(admin_data.password)
        db.commit()
        db.refresh(existing)
        return existing

    import uuid
    secret_suffix = uuid.uuid4().hex[:8].upper()
    qr_key = f"CHAKRA-{admin_data.roll_number}-{secret_suffix}"

    db_admin = Student(
        full_name=admin_data.full_name,
        college_email=admin_data.college_email,
        roll_number=admin_data.roll_number,
        phone_number=admin_data.phone_number,
        branch="ADMIN",
        year=1,
        password_hash=get_password_hash(admin_data.password),
        qr_key=qr_key,
        is_admin=True,
        admin_role="attendance"
    )
    db.add(db_admin)
    db.commit()
    db.refresh(db_admin)
    return db_admin

@router.delete("/scan-admins/{admin_id}")
def delete_scan_admin(admin_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Deletes a Scan Admin."""
    db_admin = db.query(Student).filter(Student.id == admin_id, Student.is_admin == True, Student.admin_role == "attendance").first()
    if not db_admin:
        raise HTTPException(status_code=404, detail="Scan Admin not found.")

    db.delete(db_admin)
    db.commit()
    return {"detail": "Scan Admin deleted successfully."}

# ----------------- SUPER ADMINS MANAGEMENT -----------------

@router.get("/super-admins", response_model=List[ScanAdminResponse])
def list_super_admins(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Lists all Super Admins."""
    return db.query(Student).filter(Student.is_admin == True, Student.admin_role == "super").order_by(Student.created_at.desc()).all()

@router.post("/super-admins", response_model=ScanAdminResponse)
def create_super_admin(admin_data: ScanAdminCreate, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Creates a new Super Admin or promotes/updates an existing user."""
    existing = db.query(Student).filter(
        (func.lower(Student.college_email) == func.lower(admin_data.college_email)) |
        (func.lower(Student.roll_number) == func.lower(admin_data.roll_number))
    ).first()

    if existing:
        # Promote or update existing user
        existing.is_admin = True
        existing.admin_role = "super"
        existing.full_name = admin_data.full_name
        existing.phone_number = admin_data.phone_number
        if admin_data.password:
            existing.password_hash = get_password_hash(admin_data.password)
        db.commit()
        db.refresh(existing)
        return existing

    import uuid
    secret_suffix = uuid.uuid4().hex[:8].upper()
    qr_key = f"CHAKRA-{admin_data.roll_number}-{secret_suffix}"

    db_admin = Student(
        full_name=admin_data.full_name,
        college_email=admin_data.college_email,
        roll_number=admin_data.roll_number,
        phone_number=admin_data.phone_number,
        branch="ADMIN",
        year=1,
        password_hash=get_password_hash(admin_data.password),
        qr_key=qr_key,
        is_admin=True,
        admin_role="super"
    )
    db.add(db_admin)
    db.commit()
    db.refresh(db_admin)
    return db_admin

@router.delete("/super-admins/{admin_id}")
def delete_super_admin(admin_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Deletes a Super Admin (prevents self-deletion)."""
    if admin_id == current_admin.id:
        raise HTTPException(status_code=400, detail="You cannot delete yourself.")

    db_admin = db.query(Student).filter(Student.id == admin_id, Student.is_admin == True, Student.admin_role == "super").first()
    if not db_admin:
        raise HTTPException(status_code=404, detail="Super Admin not found.")

    db.delete(db_admin)
    db.commit()
    return {"detail": "Super Admin deleted successfully."}

# ----------------- STUDENT DELETION -----------------

@router.delete("/students/{student_id}")
def delete_student(student_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Deletes a student entirely from the database."""
    student = db.query(Student).filter(Student.id == student_id, Student.is_admin == False).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found.")

    db.delete(student)
    db.commit()
    return {"detail": "Student and all their associated records deleted successfully."}

# ----------------- FEEDBACK MANAGEMENT -----------------

@router.get("/feedback")
def get_all_feedback(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Lists all student feedback submissions (Super Admin only)."""
    feedbacks = db.query(Feedback).join(Student, Feedback.student_id == Student.id).order_by(Feedback.submitted_at.desc()).all()
    result = []
    for f in feedbacks:
        result.append({
            "id": f.id,
            "student_id": f.student_id,
            "student_name": f.student.full_name,
            "student_roll": f.student.roll_number,
            "student_email": f.student.college_email,
            "student_branch": f.student.branch,
            "student_year": f.student.year,
            "q1_dsa_difficulty": f.q1_dsa_difficulty,
            "q2_dsa_clarity": f.q2_dsa_clarity,
            "q3_time_spent": f.q3_time_spent,
            "q4_solving_mode": f.q4_solving_mode,
            "q5_prompting_used": f.q5_prompting_used,
            "q6_prompting_effectiveness": f.q6_prompting_effectiveness,
            "q7_prompt_type": f.q7_prompt_type,
            "q8_prompt_challenge": f.q8_prompt_challenge,
            "q9_concept_understanding": f.q9_concept_understanding,
            "q10_platform_rating": f.q10_platform_rating,
            "q11_attendance_experience": f.q11_attendance_experience,
            "q12_codechef_interest": f.q12_codechef_interest,
            "q13_future_topics": f.q13_future_topics,
            "q14_prompting_improvement": f.q14_prompting_improvement,
            "q15_general_feedback": f.q15_general_feedback,
            "submitted_at": f.submitted_at.isoformat() + "Z"
        })
    return result

@router.get("/feedback/export")
def export_feedback(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Exports all student feedback submissions to an Excel sheet (Super Admin only)."""
    feedbacks = db.query(Feedback).join(Student, Feedback.student_id == Student.id).order_by(Student.roll_number).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Feedback"
    
    # Title Banner
    ws.merge_cells("A1:U1")
    ws["A1"] = "Chakravyuha DSA Challenge & Prompting - Student Feedback Report"
    ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="8C7030", end_color="8C7030", fill_type="solid")
    ws.row_dimensions[1].height = 40
    
    headers = [
        "Roll Number", "Full Name", "Email", "Branch", "Year",
        "Q1: Overall Event Rating", "Q2: Liked Event Structure", "Q3: Met Expectations", "Q4: Learned Anything New",
        "Q5: Improved Coding Confidence", "Q6: DSA Concept Understanding", "Q7: Practical Application Helpfulness",
        "Q8: Most Understood DSA Concepts", "Q9: AI Prompting Helpfulness", "Q10: Platform Experience Rating",
        "Q11: Problem Statement Clarity", "Q12: Future Attendance Likelihood", "Q13: Favorite Event Aspects",
        "Q14: Recommend to Peers", "Q15: Coordinator Suggestions", "Submitted At"
    ]
    ws.append([]) # blank row 2
    ws.append(headers) # Row 3
    
    header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for f in feedbacks:
        row_data = [
            f.student.roll_number,
            f.student.full_name,
            f.student.college_email,
            f.student.branch,
            f.student.year,
            f.q1_dsa_difficulty,
            f.q2_dsa_clarity,
            f.q3_time_spent,
            f.q4_solving_mode,
            f.q5_prompting_used,
            f.q6_prompting_effectiveness,
            f.q7_prompt_type,
            f.q8_prompt_challenge,
            f.q9_concept_understanding,
            f.q10_platform_rating,
            f.q11_attendance_experience,
            f.q12_codechef_interest,
            f.q13_future_topics,
            f.q14_prompting_improvement,
            f.q15_general_feedback,
            f.submitted_at.strftime("%d-%m-%Y %I:%M %p")
        ]
        ws.append(row_data)
        
    for row in range(4, ws.max_row + 1):
        row_fill = PatternFill(start_color="F9F9F9" if row % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, 22):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in [1, 4, 5, 6, 7, 8, 9, 10, 11, 14, 15, 16, 17, 19, 21]:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_feedback_report.xlsx"}
    )

# ----------------- EVENTS MANAGEMENT (SUPER ADMIN) -----------------

@router.get("/events", response_model=List[EventResponse])
def list_events_admin(current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Lists all events (Super Admin only)."""
    return db.query(Event).order_by(Event.created_at.desc()).all()

@router.post("/events", response_model=EventResponse)
def create_event(event_data: EventCreate, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Creates a new event (Super Admin only)."""
    existing = db.query(Event).filter(func.lower(Event.name) == func.lower(event_data.name)).first()
    if existing:
        raise HTTPException(status_code=400, detail="An event with this name already exists.")
        
    event = Event(
        name=event_data.name,
        description=event_data.description,
        status=event_data.status
    )
    db.add(event)
    db.commit()
    db.refresh(event)
    return event

@router.delete("/events/{event_id}")
def delete_event(event_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Deletes an event (Super Admin only)."""
    try:
        parsed_id = int(str(event_id).strip())
    except ValueError:
        raise HTTPException(status_code=400, detail="event_id must be a valid integer")
        
    event = db.query(Event).filter(Event.id == parsed_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found.")
        
    if "YUKTI" in event.name.upper():
        raise HTTPException(status_code=400, detail="Core YUKTI event cannot be deleted.")
        
    db.delete(event)
    db.commit()
    return {"detail": "Event deleted successfully."}

@router.get("/events/{event_id}/registrations")
def get_event_registrations(event_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Lists all students registered for a specific event with problem and attendance stats."""
    try:
        parsed_id = int(str(event_id).strip())
    except ValueError:
        raise HTTPException(status_code=400, detail="event_id must be a valid integer")
        
    event = db.query(Event).filter(Event.id == parsed_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found.")
        
    regs = db.query(EventRegistration).filter(EventRegistration.event_id == parsed_id).all()
    
    # Pre-fetch solved counts and attendance counts to avoid N+1 database queries
    solved_counts = dict(
        db.query(Submission.student_id, func.count(Submission.id))
        .filter(Submission.solved == True)
        .group_by(Submission.student_id).all()
    )
    
    attendance_counts = dict(
        db.query(Attendance.student_id, func.count(Attendance.id))
        .group_by(Attendance.student_id).all()
    )
    
    result = []
    for r in regs:
        student_id = r.student.id
        result.append({
            "student_id": student_id,
            "full_name": r.student.full_name,
            "college_email": r.student.college_email,
            "roll_number": r.student.roll_number or "N/A (Personal Registration)",
            "phone_number": r.student.phone_number,
            "branch": r.student.branch,
            "year": r.student.year,
            "streak_count": r.student.streak_count,
            "problems_solved": solved_counts.get(student_id, 0),
            "attendance_count": attendance_counts.get(student_id, 0),
            "registered_at": r.registered_at.isoformat() + "Z",
            "attended": r.attended,
            "attended_at": r.attended_at.isoformat() + "Z" if r.attended_at else None,
            "attendance_marked_by": r.attendance_marked_by
        })
    return {
        "event_name": event.name,
        "registrations_count": len(result),
        "students": result
    }

@router.get("/events/{event_id}/export")
def export_event_registrations(event_id: str, current_admin: Student = Depends(get_current_super_admin), db: Session = Depends(get_db)):
    """Exports registered students for a specific event to an Excel sheet."""
    try:
        parsed_id = int(str(event_id).strip())
    except ValueError:
        raise HTTPException(status_code=400, detail="event_id must be a valid integer")
        
    event = db.query(Event).filter(Event.id == parsed_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found.")
        
    regs = db.query(EventRegistration).filter(EventRegistration.event_id == parsed_id).all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Event Registrations"
    
    # Title Banner
    ws.merge_cells("A1:G1")
    ws["A1"] = f"Event Registration Report: {event.name}"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    ws.row_dimensions[1].height = 40
    
    headers = ["Full Name", "Roll Number", "College/Personal Email", "Phone Number", "Branch", "Year", "Registration Date"]
    ws.append([]) # Row 2 blank
    ws.append(headers) # Row 3
    
    header_fill = PatternFill(start_color="2A2A2A", end_color="2A2A2A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='DDDDDD'),
        right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'),
        bottom=Side(style='thin', color='DDDDDD')
    )
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    for r in regs:
        row_data = [
            r.student.full_name,
            r.student.roll_number or "N/A (Personal Registration)",
            r.student.college_email,
            r.student.phone_number,
            r.student.branch,
            r.student.year,
            r.registered_at.strftime("%d-%m-%Y %I:%M %p")
        ]
        ws.append(row_data)
        
    for row in range(4, ws.max_row + 1):
        row_fill = PatternFill(start_color="F9F9F9" if row % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in [2, 5, 6, 7]:
                cell.alignment = Alignment(horizontal="center")
                
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"Registrations_{event.name.replace(' ', '_')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/sih/teams")
def create_sih_team_admin(
    team_data: SIHTeamRegistration,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Super Admin manually creates a new SIH team (bypasses registration deadline and 1st-year cap)."""

    # 1. Check team name uniqueness
    existing_team = db.query(SIHTeam).filter(
        func.lower(SIHTeam.team_name) == func.lower(team_data.team_name.strip())
    ).first()
    if existing_team:
        raise HTTPException(
            status_code=400,
            detail=f"Team name '{team_data.team_name}' is already taken. Please choose a different name."
        )

    # 2. Check team size
    if len(team_data.members) != 5:
        raise HTTPException(
            status_code=400,
            detail="A team must consist of exactly 1 Team Leader and 5 teammates (total 6 members)."
        )

    all_members = [team_data.leader] + team_data.members

    # 3. Check college domain matching
    domains = []
    for m in all_members:
        if "@" not in m.college_email:
            raise HTTPException(status_code=400, detail=f"Invalid college email format: {m.college_email}")
        domains.append(m.college_email.lower().split("@")[1].strip())

    if len(set(domains)) > 1:
        raise HTTPException(
            status_code=400,
            detail="All team members must belong to the exact same college (identical college email domains)."
        )

    # 4. Check female member requirement
    genders = [m.gender for m in all_members]
    if "Woman" not in genders:
        raise HTTPException(
            status_code=400,
            detail="At least one female member (Woman) is mandatory in the team."
        )

    # 5. Check all members are registered on the website
    all_college_emails = [m.college_email.lower().strip() for m in all_members]
    registered_students = db.query(Student).filter(
        func.lower(Student.college_email).in_(all_college_emails)
    ).all()
    registered_emails = {s.college_email.lower().strip() for s in registered_students}

    for m in all_members:
        email = m.college_email.lower().strip()
        if email not in registered_emails:
            raise HTTPException(
                status_code=400,
                detail=f"'{m.full_name}' ({m.college_email}) is not registered on the Chakravyuha website."
            )

    all_personal_emails = [m.personal_email.lower().strip() for m in all_members]

    # 6. Check no member is already in another team
    existing_member = db.query(SIHTeamMember).filter(
        (func.lower(SIHTeamMember.college_email).in_(all_college_emails)) |
        (func.lower(SIHTeamMember.personal_email).in_(all_personal_emails)) |
        (func.lower(SIHTeamMember.college_email).in_(all_personal_emails)) |
        (func.lower(SIHTeamMember.personal_email).in_(all_college_emails))
    ).first()

    if existing_member:
        raise HTTPException(
            status_code=400,
            detail=f"'{existing_member.full_name}' ({existing_member.college_email}) is already registered in another SIH team."
        )

    # 7. Find leader's Student record for leader_student_id
    leader_student = db.query(Student).filter(
        func.lower(Student.college_email) == team_data.leader.college_email.lower().strip()
    ).first()
    if not leader_student:
        raise HTTPException(status_code=400, detail="Leader's student record not found.")

    # 8. Create team
    team = SIHTeam(
        team_name=team_data.team_name.strip(),
        leader_student_id=leader_student.id,
        room_number=team_data.room_number.strip() if team_data.room_number and team_data.room_number.strip() else None
    )
    db.add(team)
    db.commit()
    db.refresh(team)

    # 9. Save leader member
    leader_member = SIHTeamMember(
        team_id=team.id,
        is_leader=True,
        full_name=team_data.leader.full_name,
        college_email=team_data.leader.college_email.lower().strip(),
        personal_email=team_data.leader.personal_email.lower().strip(),
        phone_number=team_data.leader.phone_number,
        study_year=team_data.leader.study_year,
        branch=team_data.leader.branch,
        roll_number=team_data.leader.roll_number.upper().strip(),
        gender=team_data.leader.gender
    )
    db.add(leader_member)

    # 10. Save teammate members
    for tm in team_data.members:
        teammate = SIHTeamMember(
            team_id=team.id,
            is_leader=False,
            full_name=tm.full_name,
            college_email=tm.college_email.lower().strip(),
            personal_email=tm.personal_email.lower().strip(),
            phone_number=tm.phone_number,
            study_year=tm.study_year,
            branch=tm.branch,
            roll_number=tm.roll_number.upper().strip(),
            gender=tm.gender
        )
        db.add(teammate)

    db.commit()

    # 11. Auto-register all 6 members to the SIH event
    sih_event = db.query(Event).filter(Event.name.like("%Smart India Hackathon%")).first()
    if sih_event:
        for member_student in registered_students:
            existing_reg = db.query(EventRegistration).filter(
                EventRegistration.student_id == member_student.id,
                EventRegistration.event_id == sih_event.id
            ).first()
            if not existing_reg:
                reg = EventRegistration(
                    student_id=member_student.id,
                    event_id=sih_event.id
                )
                db.add(reg)
        db.commit()

    # 12. Send confirmation email to every team member (same as student registration)
    webhook_url = os.environ.get("POWER_AUTOMATE_SIGNUP_WEBHOOK_URL") or os.environ.get("POWER_AUTOMATE_EVENT_WEBHOOK_URL")
    if webhook_url:
        import requests

        roster_rows = ""
        for i, m in enumerate(all_members):
            role = "Leader" if i == 0 else f"Member {i}"
            roster_rows += f"""
            <tr>
              <td style="padding:6px 0;font-weight:700;color:#c5a059;">{role}</td>
              <td style="padding:6px 0;color:#ffffff;">{m.full_name} ({m.roll_number})</td>
              <td style="padding:6px 0;font-size:12px;color:#a1a1aa;">{m.college_email}</td>
            </tr>
            """

        for s in registered_students:
            qr_image_url = f"https://api.qrserver.com/v1/create-qr-code/?size=300x300&data={s.qr_key}"
            roll = s.roll_number or "N/A"

            html_body = f"""
<div style="font-family:'Segoe UI',Helvetica,Arial,sans-serif;max-width:620px;margin:0 auto;background:#0a0908;border:1px solid #c5a059;border-radius:12px;overflow:hidden;">

  <div style="background:linear-gradient(135deg,#1a1508 0%,#0a0908 50%,#1a1508 100%);padding:40px 30px 30px;text-align:center;border-bottom:2px solid #d4af37;">
    <div style="width:60px;height:60px;margin:0 auto 16px;border:2px solid #d4af37;border-radius:50%;line-height:60px;font-size:28px;">🚀</div>
    <h1 style="color:#d4af37;text-transform:uppercase;letter-spacing:4px;font-family:Georgia,serif;font-size:24px;margin:0 0 4px;">Smart India Hackathon 2026</h1>
    <p style="font-size:10px;text-transform:uppercase;color:#8c7030;letter-spacing:5px;margin:0;">Chakravyuha Club &bull; Internal Hackathon</p>
  </div>

  <div style="padding:35px 30px 10px;">
    <p style="font-size:13px;color:#d4af37;text-transform:uppercase;letter-spacing:3px;font-weight:bold;margin:0 0 12px;">🏆 Team Registration Confirmed</p>
    <h2 style="font-size:20px;color:#ffffff;margin:0 0 16px;font-weight:700;line-height:1.3;">
      Congratulations, Team <span style="color:#d4af37;">{team_data.team_name}</span>!
    </h2>
    <p style="font-size:14px;color:#d4d4d8;line-height:1.8;margin:0 0 8px;">
      Your team has been successfully registered for the Smart India Hackathon 2026 Internal Hackathon.
    </p>
    <p style="font-size:13px;color:#a1a1aa;line-height:1.7;margin:0 0 25px;">
      All team members must present their permanent attendance QR codes at the hackathon check-in desk on the day of the event.
    </p>
  </div>

  <div style="margin:0 30px 30px;background:linear-gradient(135deg,#1c1917,#151310);border:1px solid rgba(212,175,55,0.25);border-radius:10px;overflow:hidden;">
    <div style="background:rgba(212,175,55,0.08);padding:12px 20px;border-bottom:1px solid rgba(212,175,55,0.15);">
      <p style="margin:0;font-size:10px;text-transform:uppercase;letter-spacing:3px;color:#d4af37;font-weight:800;">👥 Team Roster</p>
    </div>
    <div style="padding:20px;">
      <table style="width:100%;font-size:13px;color:#e4e4e7;border-collapse:collapse;">
        {roster_rows}
      </table>
    </div>
  </div>

  <div style="text-align:center;padding:0 30px 35px;">
    <p style="font-size:11px;text-transform:uppercase;letter-spacing:3px;color:#d4af37;font-weight:700;margin:0 0 6px;">Your Credentials QR Code</p>
    <p style="font-size:12px;color:#71717a;margin:0 0 20px;">Present this at the check-in scanner</p>
    <div style="display:inline-block;padding:16px;background:#ffffff;border:3px solid #d4af37;border-radius:12px;box-shadow:0 8px 30px rgba(212,175,55,0.15);">
      <img src="{qr_image_url}" alt="Profile QR Code" style="width:180px;height:180px;display:block;" />
    </div>
  </div>

  <div style="text-align:center;padding:20px 30px 30px;">
    <p style="font-size:15px;color:#ffffff;font-weight:700;margin:0 0 6px;">Best of luck in the Hackathon! 🚀</p>
    <p style="font-size:12px;color:#71717a;margin:0 0 24px;">Build something extraordinary.</p>
    <div style="border-top:1px solid rgba(212,175,55,0.15);padding-top:20px;">
      <p style="font-size:10px;color:#52525b;text-transform:uppercase;letter-spacing:2px;margin:0;">Chakravyuha &bull; Official Coding &amp; DSA Club of Amrita &bull; Amaravati</p>
    </div>
  </div>

</div>
"""
            subject = f"Successfully registered SIH 2026 Team: {team_data.team_name}"
            payload = {
                "email": s.college_email,
                "full_name": s.full_name,
                "roll_number": roll,
                "qr_key": s.qr_key,
                "qr_image_url": qr_image_url,
                "subject": subject,
                "html_body": html_body
            }
            try:
                requests.post(webhook_url, json=payload, timeout=10)
            except Exception:
                pass

    return {"detail": f"Team '{team_data.team_name}' registered successfully!", "team_id": team.id}


@router.delete("/sih/teams/{team_id}")
def delete_sih_team(
    team_id: int,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Deletes an SIH team and removes members from event registrations (Super Admin only)."""
    team = db.query(SIHTeam).filter(SIHTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found.")
        
    db_members = db.query(SIHTeamMember).filter(SIHTeamMember.team_id == team.id).all()
    sih_event = db.query(Event).filter(Event.name.like("%Smart India Hackathon%")).first()
    if sih_event:
        college_emails = [m.college_email.lower().strip() for m in db_members]
        students = db.query(Student).filter(func.lower(Student.college_email).in_(college_emails)).all()
        student_ids = [s.id for s in students]
        
        if student_ids:
            db.query(EventRegistration).filter(
                EventRegistration.event_id == sih_event.id,
                EventRegistration.student_id.in_(student_ids)
            ).delete(synchronize_session=False)
            
    db.query(SIHTeamMember).filter(SIHTeamMember.team_id == team.id).delete()
    db.delete(team)
    db.commit()
    return {"detail": "Team deleted successfully."}

@router.put("/sih/teams/{team_id}")
def update_sih_team_admin(
    team_id: int,
    team_data: SIHTeamRegistration,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Updates an SIH team from Super Admin dashboard."""
    team = db.query(SIHTeam).filter(SIHTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found.")

    existing_team = db.query(SIHTeam).filter(
        func.lower(SIHTeam.team_name) == func.lower(team_data.team_name.strip()),
        SIHTeam.id != team.id
    ).first()
    if existing_team:
        raise HTTPException(
            status_code=400,
            detail=f"Team name '{team_data.team_name}' is already taken by another team."
        )

    if len(team_data.members) != 5:
        raise HTTPException(
            status_code=400,
            detail="A team must consist of exactly 1 Team Leader and 5 teammates (total 6 members)."
        )

    all_members = [team_data.leader] + team_data.members
    domains = []
    for m in all_members:
        if "@" not in m.college_email:
            raise HTTPException(status_code=400, detail=f"Invalid college email format: {m.college_email}")
        domains.append(m.college_email.lower().split("@")[1].strip())

    if len(set(domains)) > 1:
        raise HTTPException(
            status_code=400,
            detail="All team members must belong to the exact same college (identical college email domains)."
        )

    genders = [m.gender for m in all_members]
    if "Woman" not in genders:
        raise HTTPException(
            status_code=400,
            detail="At least one female member (Woman) is mandatory in the team to nominate for SIH 2026."
        )

    all_college_emails = [m.college_email.lower().strip() for m in all_members]
    registered_students = db.query(Student).filter(
        func.lower(Student.college_email).in_(all_college_emails)
    ).all()
    registered_emails = {s.college_email.lower().strip() for s in registered_students}

    for m in all_members:
        email = m.college_email.lower().strip()
        if email not in registered_emails:
            raise HTTPException(
                status_code=400,
                detail=f"Teammate '{m.full_name}' ({m.college_email}) is not registered on the Chakravyuha website. All team members must be registered to participate in SIH."
            )

    all_personal_emails = [m.personal_email.lower().strip() for m in all_members]

    # Query duplicate members not in current team
    existing_member = db.query(SIHTeamMember).filter(
        (SIHTeamMember.team_id != team.id) & (
            (func.lower(SIHTeamMember.college_email).in_(all_college_emails)) |
            (func.lower(SIHTeamMember.personal_email).in_(all_personal_emails)) |
            (func.lower(SIHTeamMember.college_email).in_(all_personal_emails)) |
            (func.lower(SIHTeamMember.personal_email).in_(all_college_emails))
        )
    ).first()

    if existing_member:
        raise HTTPException(
            status_code=400,
            detail=f"Member '{existing_member.full_name}' with email '{existing_member.college_email}' is already registered in another SIH team."
        )

    team.team_name = team_data.team_name.strip()
    team.room_number = team_data.room_number.strip() if team_data.room_number and team_data.room_number.strip() else None

    # Re-fetch previous members to update event registration
    db_members = db.query(SIHTeamMember).filter(SIHTeamMember.team_id == team.id).all()
    db.query(SIHTeamMember).filter(SIHTeamMember.team_id == team.id).delete()

    leader_member = SIHTeamMember(
        team_id=team.id,
        is_leader=True,
        full_name=team_data.leader.full_name,
        college_email=team_data.leader.college_email.lower().strip(),
        personal_email=team_data.leader.personal_email.lower().strip(),
        phone_number=team_data.leader.phone_number,
        study_year=team_data.leader.study_year,
        branch=team_data.leader.branch,
        roll_number=team_data.leader.roll_number.upper().strip(),
        gender=team_data.leader.gender
    )
    db.add(leader_member)

    for tm in team_data.members:
        teammate = SIHTeamMember(
            team_id=team.id,
            is_leader=False,
            full_name=tm.full_name,
            college_email=tm.college_email.lower().strip(),
            personal_email=tm.personal_email.lower().strip(),
            phone_number=tm.phone_number,
            study_year=tm.study_year,
            branch=tm.branch,
            roll_number=tm.roll_number.upper().strip(),
            gender=tm.gender
        )
        db.add(teammate)

    db.commit()

    sih_event = db.query(Event).filter(Event.name.like("%Smart India Hackathon%")).first()
    if sih_event:
        old_college_emails = [m.college_email.lower().strip() for m in db_members]
        old_students = db.query(Student).filter(
            func.lower(Student.college_email).in_(old_college_emails)
        ).all()
        old_student_ids = {s.id for s in old_students}
        
        new_student_ids = {s.id for s in registered_students}
        
        to_remove_ids = old_student_ids - new_student_ids
        if to_remove_ids:
            db.query(EventRegistration).filter(
                EventRegistration.event_id == sih_event.id,
                EventRegistration.student_id.in_(list(to_remove_ids))
            ).delete(synchronize_session=False)
            
        for member_student in registered_students:
            existing_reg = db.query(EventRegistration).filter(
                EventRegistration.student_id == member_student.id,
                EventRegistration.event_id == sih_event.id
            ).first()
            if not existing_reg:
                reg = EventRegistration(
                    student_id=member_student.id,
                    event_id=sih_event.id
                )
                db.add(reg)
        db.commit()

    return {"detail": "Team details updated successfully!"}

@router.get("/sih/teams")
def get_sih_teams(
    current_admin: Student = Depends(get_current_super_admin),
    room: str = None,
    db: Session = Depends(get_db)
):
    """Lists all registered SIH teams and member details (Super Admin only). Optionally filter by room number."""
    query = db.query(SIHTeam).order_by(SIHTeam.created_at.desc())
    if room:
        query = query.filter(func.lower(SIHTeam.room_number) == func.lower(room.strip()))
    teams = query.all()

    result = []
    for t in teams:
        members_data = []
        for m in t.members:
            members_data.append({
                "is_leader": m.is_leader,
                "full_name": m.full_name,
                "college_email": m.college_email,
                "personal_email": m.personal_email,
                "phone_number": m.phone_number,
                "study_year": m.study_year,
                "branch": m.branch,
                "roll_number": m.roll_number,
                "gender": m.gender
            })

        # Try to find the leader name in Student table
        leader_name = t.leader.full_name if t.leader else "Unknown"

        result.append({
            "id": t.id,
            "team_name": t.team_name,
            "created_at": t.created_at.isoformat() + "Z",
            "leader_student_id": t.leader_student_id,
            "leader_name": leader_name,
            "room_number": t.room_number or None,
            "members": members_data
        })

    return result


@router.get("/sih/analytics")
def get_sih_analytics(
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Computes SIH 2026 registration analytics (Super Admin only)."""
    total_teams = db.query(SIHTeam).count()
    total_students = db.query(SIHTeamMember).count()
    
    women_count = db.query(SIHTeamMember).filter(SIHTeamMember.gender == "Woman").count()
    men_count = db.query(SIHTeamMember).filter(SIHTeamMember.gender == "Man").count()
    
    branch_counts = dict(
        db.query(SIHTeamMember.branch, func.count(SIHTeamMember.id))
        .group_by(SIHTeamMember.branch).all()
    )
    
    year_counts = dict(
        db.query(SIHTeamMember.study_year, func.count(SIHTeamMember.id))
        .group_by(SIHTeamMember.study_year).all()
    )
    
    # Count all-first-year teams (all 6 members are year 1)
    non_first_year_team_ids = db.query(SIHTeamMember.team_id).filter(
        SIHTeamMember.study_year != 1
    ).distinct().all()
    excluded_ids = [row[0] for row in non_first_year_team_ids]
    all_first_year_teams = total_teams - len(excluded_ids)

    return {
        "total_teams": total_teams,
        "total_students": total_students,
        "all_first_year_teams": all_first_year_teams,
        "all_first_year_teams_cap": 20,
        "gender_breakdown": {
            "Woman": women_count,
            "Man": men_count
        },
        "branch_breakdown": branch_counts,
        "year_breakdown": {str(k): v for k, v in year_counts.items()}
    }


@router.get("/sih/export")
def export_sih_registrations(
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Exports SIH registered teams and rosters to an Excel sheet (Super Admin only)."""
    teams = db.query(SIHTeam).order_by(SIHTeam.created_at.desc()).all()

    # Build a lookup: team_id -> (ps_number, ps_title)
    ps_lookup: dict = {}
    selections = (
        db.query(SIHPSSelection, SIHProblemStatement)
        .join(SIHProblemStatement, SIHPSSelection.problem_statement_id == SIHProblemStatement.id)
        .all()
    )
    for sel, ps in selections:
        ps_lookup[sel.team_id] = (ps.ps_number, ps.title)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIH 2026 Teams"

    # Title Banner — now spans 13 columns (A–M)
    ws.merge_cells("A1:M1")
    ws["A1"] = "Smart India Hackathon 2026 — Official Team Registration Roster"
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    ws.row_dimensions[1].height = 45

    headers = [
        "Team Name", "Role", "Full Name", "Roll Number",
        "College Email", "Personal Email", "Phone Number",
        "Study Year", "Branch", "Gender", "Registration Date",
        "PS Number", "PS Title"
    ]
    ws.append([])       # Row 2 blank
    ws.append(headers)  # Row 3

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[3].height = 25

    row_num = 4
    for t in teams:
        # Get PS info for this team (empty strings if not yet selected)
        ps_number, ps_title = ps_lookup.get(t.id, ("—", "Not Selected"))

        # Sort members so leader is first, then members
        sorted_members = sorted(t.members, key=lambda x: not x.is_leader)

        for idx, m in enumerate(sorted_members):
            role_str = "Team Leader" if m.is_leader else f"Member {idx}"
            row_data = [
                t.team_name,
                role_str,
                m.full_name,
                m.roll_number,
                m.college_email,
                m.personal_email,
                m.phone_number,
                f"Year {m.study_year}",
                m.branch,
                m.gender,
                t.created_at.strftime("%d-%m-%Y %I:%M %p"),
                ps_number,
                ps_title
            ]
            ws.append(row_data)

            # Format row
            row_fill = PatternFill(start_color="F2F4F4" if row_num % 2 == 0 else "FFFFFF", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.fill = row_fill
                cell.border = thin_border

                # Alignments — centre for cols that are short/categorical
                if col in [1, 2, 4, 8, 9, 10, 11, 12]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            row_num += 1

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=SIH_2026_Teams_Roster.xlsx"}
    )


# ── Admin SIH Problem Statement Endpoints ──────────────────────────────────────

@router.get("/sih/ps-analytics")
def get_sih_ps_analytics(
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Returns PS selection analytics for the admin dashboard."""
    total_teams = db.query(SIHTeam).count()
    confirmed_team_ids = db.query(SIHPSSelection.team_id).all()
    confirmed_count = len(confirmed_team_ids)
    not_confirmed_count = total_teams - confirmed_count

    # Count by PS (how many teams chose each PS)
    ps_counts_raw = (
        db.query(SIHProblemStatement.ps_number, SIHProblemStatement.title, func.count(SIHPSSelection.id))
        .join(SIHPSSelection, SIHPSSelection.problem_statement_id == SIHProblemStatement.id)
        .group_by(SIHProblemStatement.id, SIHProblemStatement.ps_number, SIHProblemStatement.title)
        .order_by(func.count(SIHPSSelection.id).desc())
        .all()
    )

    ps_distribution = [
        {"ps_number": row[0], "title": row[1], "team_count": row[2]}
        for row in ps_counts_raw
    ]

    return {
        "total_teams": total_teams,
        "confirmed": confirmed_count,
        "not_confirmed": not_confirmed_count,
        "ps_distribution": ps_distribution,
    }


@router.get("/sih/teams-with-ps")
def get_sih_teams_with_ps(
    ps_filter: str = "all",   # "all" | "confirmed" | "not_confirmed"
    search: str = "",
    page: int = 1,
    limit: int = 20,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Returns paginated SIH teams with PS selection info and filter support."""
    query = db.query(SIHTeam)
    if search.strip():
        query = query.filter(func.lower(SIHTeam.team_name).like(f"%{search.strip().lower()}%"))

    confirmed_team_ids_set = {row[0] for row in db.query(SIHPSSelection.team_id).all()}

    if ps_filter == "confirmed":
        query = query.filter(SIHTeam.id.in_(confirmed_team_ids_set))
    elif ps_filter == "not_confirmed":
        query = query.filter(~SIHTeam.id.in_(confirmed_team_ids_set))

    total = query.count()
    offset = (page - 1) * limit
    teams = query.order_by(SIHTeam.created_at.desc()).offset(offset).limit(limit).all()

    # Build PS selection map
    sel_map = {}
    for sel in db.query(SIHPSSelection).all():
        sel_map[sel.team_id] = sel

    result = []
    for t in teams:
        sel = sel_map.get(t.id)
        ps_info = None
        if sel:
            ps = sel.problem_statement
            ps_info = {
                "ps_number": ps.ps_number,
                "ps_id": ps.ps_id,
                "title": ps.title,
                "organization": ps.organization,
                "theme": ps.theme,
                "selected_at": sel.selected_at.isoformat() + "Z",
                "last_edited_by_admin": sel.last_edited_by_admin,
            }
        # Find leader name
        leader_m = next((m for m in t.members if m.is_leader), None)
        leader_name = leader_m.full_name if leader_m else "Unknown"

        result.append({
            "id": t.id,
            "team_name": t.team_name,
            "leader_name": leader_name,
            "created_at": t.created_at.isoformat() + "Z",
            "ps_selection": ps_info,
        })

    return {"total": total, "page": page, "limit": limit, "items": result}


@router.put("/sih/teams/{team_id}/ps")
def admin_override_ps(
    team_id: int,
    payload: AdminPSOverrideRequest,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Admin overrides a team's PS selection (can edit even after confirmation)."""
    team = db.query(SIHTeam).filter(SIHTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found.")

    ps = db.query(SIHProblemStatement).filter(SIHProblemStatement.id == payload.problem_statement_id).first()
    if not ps:
        raise HTTPException(status_code=404, detail="Problem Statement not found.")

    existing = db.query(SIHPSSelection).filter(SIHPSSelection.team_id == team_id).first()
    if existing:
        existing.problem_statement_id = ps.id
        existing.selected_at = datetime.datetime.utcnow()
        existing.last_edited_by_admin = True
    else:
        selection = SIHPSSelection(
            team_id=team_id,
            problem_statement_id=ps.id,
            selected_at=datetime.datetime.utcnow(),
            last_edited_by_admin=True
        )
        db.add(selection)

    db.commit()
    return {"detail": f"PS updated to {ps.ps_number} for team '{team.team_name}' by admin."}


@router.delete("/sih/teams/{team_id}/ps")
def delete_team_ps(
    team_id: int,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Admin deletes/removes a team's PS selection (resetting it to not selected)."""
    selection = db.query(SIHPSSelection).filter(SIHPSSelection.team_id == team_id).first()
    if not selection:
        raise HTTPException(status_code=404, detail="No PS selection found for this team.")
    db.delete(selection)
    db.commit()
    return {"detail": "PS selection removed successfully."}


@router.get("/sih/ps-list")
def admin_get_ps_list(
    search: str = "",
    page: int = 1,
    limit: int = 50,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Admin-accessible full PS list for override dropdowns."""
    query = db.query(SIHProblemStatement)
    if search.strip():
        s = f"%{search.strip().lower()}%"
        query = query.filter(
            func.lower(SIHProblemStatement.ps_number).like(s) |
            func.lower(SIHProblemStatement.title).like(s)
        )
    total = query.count()
    items = query.order_by(SIHProblemStatement.ps_id).offset((page - 1) * limit).limit(limit).all()
    return {
        "total": total,
        "items": [{"id": ps.id, "ps_number": ps.ps_number, "title": ps.title[:100]} for ps in items]
    }


# ── SIH Judging / Marks Endpoints ────────────────────────────────────────────

def _raw_total(s: SIHJudgingScore) -> int:
    return (
        s.j1_problem_understanding + s.j1_innovation + s.j1_technical_feasibility +
        s.j1_scalability_impact + s.j1_presentation_qa +
        s.j2_problem_understanding + s.j2_innovation + s.j2_technical_feasibility +
        s.j2_scalability_impact + s.j2_presentation_qa
    )

def _build_score_dict(s: SIHJudgingScore, team: SIHTeam, norm_score: float) -> dict:
    raw = _raw_total(s)
    return {
        "team_id": team.id,
        "team_name": team.team_name,
        "room_number": team.room_number or "—",
        "j1_problem_understanding": s.j1_problem_understanding,
        "j1_innovation": s.j1_innovation,
        "j1_technical_feasibility": s.j1_technical_feasibility,
        "j1_scalability_impact": s.j1_scalability_impact,
        "j1_presentation_qa": s.j1_presentation_qa,
        "j2_problem_understanding": s.j2_problem_understanding,
        "j2_innovation": s.j2_innovation,
        "j2_technical_feasibility": s.j2_technical_feasibility,
        "j2_scalability_impact": s.j2_scalability_impact,
        "j2_presentation_qa": s.j2_presentation_qa,
        "criterion_problem_understanding": s.j1_problem_understanding + s.j2_problem_understanding,
        "criterion_innovation": s.j1_innovation + s.j2_innovation,
        "criterion_technical_feasibility": s.j1_technical_feasibility + s.j2_technical_feasibility,
        "criterion_scalability_impact": s.j1_scalability_impact + s.j2_scalability_impact,
        "criterion_presentation_qa": s.j1_presentation_qa + s.j2_presentation_qa,
        "raw_total": raw,
        "normalized_score": round(norm_score, 2),
        "entered_by": s.entered_by,
        "entered_at": s.entered_at.isoformat() + "Z",
        "updated_at": s.updated_at.isoformat() + "Z",
    }


def _normalize_scores(scores: list) -> list:
    """Apply per-room normalization. Returns list of (SIHJudgingScore, SIHTeam, norm_score)."""
    # Group by room_number
    from collections import defaultdict
    room_groups: dict = defaultdict(list)
    for s, t in scores:
        room_groups[t.room_number or "UNASSIGNED"].append((s, t))

    result = []
    for room, entries in room_groups.items():
        raws = [_raw_total(s) for s, _ in entries]
        room_max = max(raws) if raws else 1
        for (s, t), raw in zip(entries, raws):
            norm = (raw / room_max * 100) if room_max > 0 else 0.0
            result.append((s, t, norm))
    return result


@router.get("/sih/marks")
def get_sih_marks(
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Returns all judging scores with per-room normalized scores, sorted by normalized score desc."""
    rows = (
        db.query(SIHJudgingScore, SIHTeam)
        .join(SIHTeam, SIHJudgingScore.team_id == SIHTeam.id)
        .all()
    )
    normalized = _normalize_scores(rows)
    normalized.sort(key=lambda x: x[2], reverse=True)
    return [_build_score_dict(s, t, n) for s, t, n in normalized]


@router.get("/sih/marks/{team_id}")
def get_sih_marks_team(
    team_id: int,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Returns the judging score for a specific team with its normalized score."""
    team = db.query(SIHTeam).filter(SIHTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found.")
    score = db.query(SIHJudgingScore).filter(SIHJudgingScore.team_id == team_id).first()
    if not score:
        return None

    # Normalize within room
    room_scores = (
        db.query(SIHJudgingScore, SIHTeam)
        .join(SIHTeam, SIHJudgingScore.team_id == SIHTeam.id)
        .filter(func.coalesce(SIHTeam.room_number, "UNASSIGNED") == func.coalesce(team.room_number, "UNASSIGNED"))
        .all()
    )
    raws = [_raw_total(s) for s, _ in room_scores]
    room_max = max(raws) if raws else 1
    norm = (_raw_total(score) / room_max * 100) if room_max > 0 else 0.0
    return _build_score_dict(score, team, norm)


@router.post("/sih/marks")
def upsert_sih_marks(
    payload: dict,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Create or update judging scores for a team. Payload: team_id + 10 score fields (0-10 each)."""
    team_id = payload.get("team_id")
    if not team_id:
        raise HTTPException(status_code=400, detail="team_id is required.")
    team = db.query(SIHTeam).filter(SIHTeam.id == team_id).first()
    if not team:
        raise HTTPException(status_code=404, detail="Team not found.")

    fields = [
        "j1_problem_understanding", "j1_innovation", "j1_technical_feasibility",
        "j1_scalability_impact", "j1_presentation_qa",
        "j2_problem_understanding", "j2_innovation", "j2_technical_feasibility",
        "j2_scalability_impact", "j2_presentation_qa",
    ]
    for f in fields:
        val = payload.get(f, 0)
        if not isinstance(val, int) or val < 0 or val > 10:
            raise HTTPException(status_code=400, detail=f"'{f}' must be an integer 0–10.")

    score = db.query(SIHJudgingScore).filter(SIHJudgingScore.team_id == team_id).first()
    if score:
        for f in fields:
            setattr(score, f, payload.get(f, 0))
        score.entered_by = current_admin.college_email
        score.updated_at = datetime.datetime.utcnow()
    else:
        score = SIHJudgingScore(
            team_id=team_id,
            entered_by=current_admin.college_email,
            **{f: payload.get(f, 0) for f in fields}
        )
        db.add(score)
    db.commit()
    db.refresh(score)

    # Return with normalized score
    room_scores = (
        db.query(SIHJudgingScore, SIHTeam)
        .join(SIHTeam, SIHJudgingScore.team_id == SIHTeam.id)
        .filter(func.coalesce(SIHTeam.room_number, "UNASSIGNED") == func.coalesce(team.room_number, "UNASSIGNED"))
        .all()
    )
    raws = [_raw_total(s) for s, _ in room_scores]
    room_max = max(raws) if raws else 1
    norm = (_raw_total(score) / room_max * 100) if room_max > 0 else 0.0
    return _build_score_dict(score, team, norm)


@router.delete("/sih/marks/{team_id}")
def delete_sih_marks(
    team_id: int,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Delete judging scores for a team."""
    db.delete(score)
    db.commit()
    return {"detail": "Marks deleted successfully."}


# ── Export SIH Judging sheet ──────────────────────────────────────────────────
@router.get("/sih/export-judging")
def export_sih_judging(
    room: Optional[str] = None,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Exports SIH judging sheets with raw and normalized marks, split by judge (Super Admin only)."""
    query = db.query(SIHTeam)
    if room and room.strip() and room.strip() != "All":
        query = query.filter(func.lower(SIHTeam.room_number) == func.lower(room.strip()))
    teams = query.all()

    teams.sort(key=lambda t: (t.room_number or "ZZZ", t.team_name.lower()))

    ps_lookup: dict = {}
    selections = (
        db.query(SIHPSSelection, SIHProblemStatement)
        .join(SIHProblemStatement, SIHPSSelection.problem_statement_id == SIHProblemStatement.id)
        .all()
    )
    for sel, ps in selections:
        ps_lookup[sel.team_id] = (ps.ps_number, ps.title)

    scores_query = db.query(SIHJudgingScore, SIHTeam).join(SIHTeam, SIHJudgingScore.team_id == SIHTeam.id).all()
    normalized_list = _normalize_scores(scores_query)
    scores_lookup = {s.team_id: (s, n) for s, _, n in normalized_list}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIH Judging Sheet"

    ws.merge_cells("A1:W1")
    title_text = "Smart India Hackathon 2026 — Team Evaluation & Judging Sheet"
    if room and room.strip() and room.strip() != "All":
        title_text += f" (Room: {room.upper()})"
    ws["A1"] = title_text
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    ws.row_dimensions[1].height = 45

    headers = [
        "Room No.", "Judge 1 Name", "Judge 2 Name", "Team Name", "PS Number", "PS Title",
        "J1 - Prob. Und. (/10)", "J2 - Prob. Und. (/10)", "Total Prob. Und. (/20)",
        "J1 - Innov. (/10)", "J2 - Innov. (/10)", "Total Innov. (/20)",
        "J1 - Tech. Feas. (/10)", "J2 - Tech. Feas. (/10)", "Total Tech. Feas. (/20)",
        "J1 - Scale. (/10)", "J2 - Scale. (/10)", "Total Scale. (/20)",
        "J1 - Pres. (/10)", "J2 - Pres. (/10)", "Total Pres. (/20)",
        "Raw Total (/100)", "Normalized Score"
    ]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    ws.row_dimensions[3].height = 35

    JUDGES_MAP = {
        "A009": ("Dr. Balaji I K", "Mr. Goradla Bharadwaja"),
        "A010": ("Dr. R Satya Rajendra Singh", "Dr. Deepak Kumar Panda"),
        "A011": ("Dr. Korrapati Sindhu", "Dr. Gowthami Ummadisetti"),
        "A012": ("Dr. Kurmala Gowri Raghavendra Narayan", "Mr. Cherukuri Naresh"),
        "A013": ("Dr. K.Ashwini", "Dr. Rashmi Prasad"),
        "A109": ("Dr. Kamepalli S L Prasanna", "Mr. Kistam Gopi"),
        "A110": ("Dr. Madhusudana Rao Nalluri", "M/s. Lavanya Goura"),
        "A111": ("Dr. P N S B S V Prasad V", "M/s. Lakshmi Prasanna M"),
        "A112": ("Dr. Parvathaneni Naga Srinivasu", "Dr. Anand Kumar"),
        "A113": ("Dr. Gampa V P Chandra Sekhar Yadav", "Dr. Mohit Kumar Singh"),
        "A210": ("Dr. Ravishankar Prakash Desai", "Dr. Kesavulu"),
        "A211": ("Dr. Snehamoy Pramanik", "M/s. Venkata Anusha Kolluru"),
        "A212": ("Dr. SATHIES T", "M/s Yalamala Baby Sushma"),
        "A213": ("Dr. Thunugala Bala Krishna", "Mr. Budati Jaya Lakshmi Narayana"),
        "A310": ("Dr. Mrinalini Bhagawati", "Mr. Dontha MadhuSudhana Rao"),
        "A311": ("Dr. Suresh Reddy", "Dr. Gayathri Parasa"),
        "A312": ("Dr. Vikas Rohil", "Mr. Kamjula Lakshmi Kanth Reddy"),
        "A313": ("Dr. Ravi Sankar Puppala", "Dr. Chaitanya Bathina")
    }

    row_num = 4
    for t in teams:
        ps_number, ps_title = ps_lookup.get(t.id, ("—", "Not Selected"))
        j1_name, j2_name = JUDGES_MAP.get(t.room_number or "", ("—", "—"))

        score_info = scores_lookup.get(t.id)
        if score_info:
            s, norm = score_info
            row_data = [
                t.room_number or "—",
                j1_name,
                j2_name,
                t.team_name,
                ps_number,
                ps_title,
                s.j1_problem_understanding,
                s.j2_problem_understanding,
                s.j1_problem_understanding + s.j2_problem_understanding,
                s.j1_innovation,
                s.j2_innovation,
                s.j1_innovation + s.j2_innovation,
                s.j1_technical_feasibility,
                s.j2_technical_feasibility,
                s.j1_technical_feasibility + s.j2_technical_feasibility,
                s.j1_scalability_impact,
                s.j2_scalability_impact,
                s.j1_scalability_impact + s.j2_scalability_impact,
                s.j1_presentation_qa,
                s.j2_presentation_qa,
                s.j1_presentation_qa + s.j2_presentation_qa,
                _raw_total(s),
                round(norm, 2)
            ]
        else:
            row_data = [
                t.room_number or "—",
                j1_name,
                j2_name,
                t.team_name,
                ps_number,
                ps_title,
                0, 0, 0,
                0, 0, 0,
                0, 0, 0,
                0, 0, 0,
                0, 0, 0,
                0, 0.0
            ]

        ws.append(row_data)

        row_fill = PatternFill(start_color="F2F4F4" if row_num % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in [1, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"SIH_Judging_Sheet_{room.upper() if room else 'All'}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ── Export SIH Room Allocations ───────────────────────────────────────────────
@router.get("/sih/export-rooms")
def export_sih_rooms(
    room: Optional[str] = None,
    current_admin: Student = Depends(get_current_super_admin),
    db: Session = Depends(get_db)
):
    """Exports SIH room allocations including judge names and team member details (Super Admin only)."""
    query = db.query(SIHTeam)
    if room and room.strip() and room.strip() != "All":
        query = query.filter(func.lower(SIHTeam.room_number) == func.lower(room.strip()))
    teams = query.all()

    teams.sort(key=lambda t: (t.room_number or "ZZZ", t.team_name.lower()))

    ps_lookup: dict = {}
    selections = (
        db.query(SIHPSSelection, SIHProblemStatement)
        .join(SIHProblemStatement, SIHPSSelection.problem_statement_id == SIHProblemStatement.id)
        .all()
    )
    for sel, ps in selections:
        ps_lookup[sel.team_id] = (ps.ps_number, ps.title)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SIH Room Allocations"

    ws.merge_cells("A1:K1")
    title_text = "Smart India Hackathon 2026 — Team Room Allocations & Judges"
    if room and room.strip() and room.strip() != "All":
        title_text += f" (Room: {room.upper()})"
    ws["A1"] = title_text
    ws["A1"].font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
    ws.row_dimensions[1].height = 45

    headers = [
        "S.No", "Room No.", "Team Name", "Leader Name", "Leader Roll",
        "Leader Email", "Teammates Details", "Judge 1 Name", "Judge 2 Name",
        "PS Number", "PS Title"
    ]
    ws.append([])
    ws.append(headers)

    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    ws.row_dimensions[3].height = 25

    JUDGES_MAP = {
        "A009": ("Dr. Balaji I K", "Mr. Goradla Bharadwaja"),
        "A010": ("Dr. R Satya Rajendra Singh", "Dr. Deepak Kumar Panda"),
        "A011": ("Dr. Korrapati Sindhu", "Dr. Gowthami Ummadisetti"),
        "A012": ("Dr. Kurmala Gowri Raghavendra Narayan", "Mr. Cherukuri Naresh"),
        "A013": ("Dr. K.Ashwini", "Dr. Rashmi Prasad"),
        "A109": ("Dr. Kamepalli S L Prasanna", "Mr. Kistam Gopi"),
        "A110": ("Dr. Madhusudana Rao Nalluri", "M/s. Lavanya Goura"),
        "A111": ("Dr. P N S B S V Prasad V", "M/s. Lakshmi Prasanna M"),
        "A112": ("Dr. Parvathaneni Naga Srinivasu", "Dr. Anand Kumar"),
        "A113": ("Dr. Gampa V P Chandra Sekhar Yadav", "Dr. Mohit Kumar Singh"),
        "A210": ("Dr. Ravishankar Prakash Desai", "Dr. Kesavulu"),
        "A211": ("Dr. Snehamoy Pramanik", "M/s. Venkata Anusha Kolluru"),
        "A212": ("Dr. SATHIES T", "M/s Yalamala Baby Sushma"),
        "A213": ("Dr. Thunugala Bala Krishna", "Mr. Budati Jaya Lakshmi Narayana"),
        "A310": ("Dr. Mrinalini Bhagawati", "Mr. Dontha MadhuSudhana Rao"),
        "A311": ("Dr. Suresh Reddy", "Dr. Gayathri Parasa"),
        "A312": ("Dr. Vikas Rohil", "Mr. Kamjula Lakshmi Kanth Reddy"),
        "A313": ("Dr. Ravi Sankar Puppala", "Dr. Chaitanya Bathina")
    }

    row_num = 4
    for idx, t in enumerate(teams, 1):
        ps_number, ps_title = ps_lookup.get(t.id, ("—", "Not Selected"))
        j1_name, j2_name = JUDGES_MAP.get(t.room_number or "", ("—", "—"))

        leader_m = next((m for m in t.members if m.is_leader), None)
        teammates = [m for m in t.members if not m.is_leader]
        
        leader_name = leader_m.full_name if leader_m else "—"
        leader_roll = leader_m.roll_number if leader_m else "—"
        leader_email = leader_m.college_email if leader_m else "—"

        teammates_str = ", ".join([f"{m.full_name} ({m.roll_number})" for m in teammates])

        row_data = [
            idx,
            t.room_number or "—",
            t.team_name,
            leader_name,
            leader_roll,
            leader_email,
            teammates_str,
            j1_name,
            j2_name,
            ps_number,
            ps_title
        ]
        ws.append(row_data)

        row_fill = PatternFill(start_color="F2F4F4" if row_num % 2 == 0 else "FFFFFF", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            if col in [1, 2, 5, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        row_num += 1

    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            if cell.row == 1:
                continue
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        if col_letter == 'G':
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    filename = f"SIH_Room_Allocations_{room.upper() if room else 'All'}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

